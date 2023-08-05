
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
  
book = openpyxl.load_workbook('D:\\attendance.xlsx')
  
sheet = book['Sheet1']
  
# counting number of rows / students
r = sheet.max_row
  
# variable for looping for input
resp = 1
  
# counting number of columns / subjects
c = sheet.max_column
  
# list of students to remind
l1 = []
  
# to concatenate list of roll numbers with
# lack of attendance
l2 = ""
  
# list of roll numbers with lack of attendance
l3 = []
  
# staff mail ids
staff_mails = ['erakshaya485@gmail.com', 'yyyyyyyy@gmail.com']
  
# Warning messages
m1 = "warning!!! you can take only one more day leave for CI class"
m2 = "warning!!! you can take only one more day leave for python class"
m3 = "warning!!! you can take only one more day leave for DM class"
  
  
def savefile():
    book.save(r'D:\\attendance.xlsx')
    print("saved!")
  
  
def check(no_of_days, row_num, b):
  
    # to use the globally declared lists and strings
    global staff_mails
    global l2
    global l3
  
    for student in range(0, len(row_num)):
        # if total no.of.leaves equals threshold
        if no_of_days[student] is 2:
            if b is 1:
                  
                # mail_id appending
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m1)  # sending mail
            elif b is 2:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m2)
            else:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m3)
  
        # if total.no.of.leaves > threshold
        elif no_of_days[student] > 2:
            if b is 1:
  
                # adding roll no
                l2 = l2+str(sheet.cell(row=row_num[student], column=1).value)
  
                # student mail_id appending
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "CI"  # subject based on the code number
  
            elif b is 2:
                l2 = l2+str(sheet.cell(row=row_num[student], column=1).value)
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "Python"
  
            else:
                l2 = l2+str(sheet.cell(row=row_num[student], column=1).value)
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "Data mining"
  
        # If threshold crossed, modify the message
        if l2 != "" and len(l3) != 0:
  
            # message for student
            msg1 = "you have lack of attendance in " + subject + " !!!"
  
            # message for staff
            msg2 = "the following students have lack of attendance in your subject : "+l2
  
            mailstu(l3, msg1)  # mail to students
            staff_id = staff_mails[b-1]  # pick respective staff's mail_id
            mailstaff(staff_id, msg2)  # mail to staff
  
# for students
def mailstu(li, msg):
    from_id = 'crazygirlaks@gmail.com'
    pwd = 'ERAkshaya485'
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
  
    # for each student to warn send mail
    for i in range(0, len(li)):
        to_id = li[i]
        message = MIMEMultipart()
        message['Subject'] = 'Attendance report'
        message.attach(MIMEText(msg, 'plain'))
        content = message.as_string()
        s.sendmail(from_id, to_id, content)
        s.quit()
    print("mail sent to students")
  
# for staff
def mailstaff(mail_id, msg):
    from_id = 'crazygirlaks@gmail.com'
    pwd = 'ERAkshaya485'
    to_id = mail_id
    message = MIMEMultipart()
    message['Subject'] = 'Lack of attendance report'
    message.attach(MIMEText(msg, 'plain'))
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
    content = message.as_string()
    s.sendmail(from_id, to_id, content)
    s.quit()
    print('Mail Sent to staff')
  
  
while resp is 1:
    print("1--->CI\n2--->python\n3--->DM")
  
    # enter the correspondingnumber
    y = int(input("enter subject :"))
  
    # no.of.absentees for that subject
    no_of_absentees = int(input('no.of.absentees :'))
  
    if(no_of_absentees > 1):
        x = list(map(int, (input('roll nos :').split(' '))))
    else:
        x = [int(input('roll no :'))]
  
    # list to hold row of the student in Excel sheet
    row_num = []
  
    # list to hold total no.of leaves
    # taken by ith student
    no_of_days = []
  
    for student in x:
  
        for i in range(2, r+1):
  
            if y is 1:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=3).value
                    m = m+1
                    sheet.cell(row=i, column=3).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
  
            elif y is 2:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=4).value
                    m = m+1
                    sheet.cell(row=i, column=4).value = m+1
                    no_of_days.append(m)
                    row_num.append(i)
  
            elif y is 3:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=5).value
                    m = m+1
                    sheet.cell(row=i, column=5).value = m+1
                    row_num.append(i)
                    no_of_days.append(m)
  
    check(no_of_days, row_num, y)
    resp = int(input('another subject ? 1---->yes 0--->no'))
